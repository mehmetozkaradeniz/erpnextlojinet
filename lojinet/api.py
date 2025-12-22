# -*- coding: utf-8 -*-
# Copyright (c) 2025, Lojinet Team
# For license information, please see license.txt

from __future__ import unicode_literals
import frappe
from frappe import _
from frappe.utils import flt, today, now, add_days

# ============== FİYAT ANLAŞMASI İŞLEMLERİ ==============
@frappe.whitelist()
def get_price_agreement(customer, date=None):
    """Müşteri için geçerli fiyat anlaşmasını getir"""
    if not date:
        date = today()
    
    # Geçerli fiyat anlaşmasını bul
    agreement = frappe.db.sql("""
        SELECT fa.name
        FROM `tabLojinet Fiyat Anlasmasi` fa
        WHERE fa.cari = %s
        AND (fa.baslangic_tarihi IS NULL OR fa.baslangic_tarihi <= %s)
        AND (fa.bitis_tarihi IS NULL OR fa.bitis_tarihi >= %s)
        AND fa.docstatus != 2
        ORDER BY fa.baslangic_tarihi DESC
        LIMIT 1
    """, (customer, date, date), as_dict=True)
    
    if not agreement:
        return []
    
    # Fiyat detaylarını getir
    details = frappe.get_all(
        "Lojinet Fiyat Anlasmasi Detay",
        filters={"parent": agreement[0].name},
        fields=["masraf_kalemi", "birim_fiyat"],
        order_by="idx"
    )
    
    return details

# ============== CARİ İŞLEMLERİ ==============
@frappe.whitelist()
def validate_customer(doc, method):
    """Customer validasyonu"""
    pass

# ============== B2B İŞLEMLERİ ==============
@frappe.whitelist()
def create_b2b_user(customer, email, first_name, last_name):
    """B2B kullanıcı oluştur"""
    if frappe.db.exists("User", email):
        frappe.throw(_("Bu email adresi zaten kayıtlı"))
    
    user = frappe.get_doc({
        "doctype": "User",
        "email": email,
        "first_name": first_name,
        "last_name": last_name,
        "send_welcome_email": 1,
        "user_type": "Website User"
    })
    user.insert(ignore_permissions=True)
    user.add_roles("Lojinet B2B User")
    frappe.db.set_value("User", user.name, "lojinet_customer", customer)
    return user.name

# ============== MUTABAKAT İŞLEMLERİ ==============
@frappe.whitelist()
def send_mutabakat(customer, month, year):
    """Online mutabakat gönder"""
    mutabakat = frappe.get_doc({
        "doctype": "Lojinet Online Mutabakat",
        "cari": customer,
        "ay": month,
        "yil": year,
        "durum": "Cevap Bekleniyor"
    })
    mutabakat.insert()
    
    cari = frappe.get_doc("Lojinet Cari", customer)
    if not cari.email:
        frappe.throw(_("Cari için email adresi tanımlı değil"))
    
    link = f"{frappe.utils.get_url()}/mutabakat/{mutabakat.name}"
    
    frappe.sendmail(
        recipients=[cari.email],
        subject=f"Online Mutabakat - {month}/{year}",
        message=f"""
        <p>Sayın {cari.cari_adi},</p>
        <p>{month}/{year} dönemi için online mutabakatınız hazırlanmıştır.</p>
        <p>Mutabakatı görüntülemek ve onaylamak için lütfen aşağıdaki linke tıklayın:</p>
        <p><a href="{link}">{link}</a></p>
        """
    )
    return mutabakat.name

@frappe.whitelist(allow_guest=True)
def onay_mutabakat(mutabakat_name, onay, aciklama=None):
    """Mutabakat onay/red"""
    import frappe.sessions
    
    mutabakat = frappe.get_doc("Lojinet Online Mutabakat", mutabakat_name)
    mutabakat.durum = "Onaylandı" if onay else "Reddedildi"
    mutabakat.onay_tarihi = now()
    mutabakat.ip_adresi = frappe.local.request_ip
    if aciklama:
        mutabakat.red_aciklamasi = aciklama
    mutabakat.save(ignore_permissions=True)
    return True

# ============== YÜK İŞLEMLERİ ==============
@frappe.whitelist()
def bulk_import_items(yuk_name, items_json):
    """Excel ile toplu ürün ekleme"""
    import json
    yuk = frappe.get_doc("Lojinet Yuk", yuk_name)
    items = json.loads(items_json)
    
    for item in items:
        yuk.append("yuk_kalemleri", {
            "stok_kodu": item.get("stok_kodu"),
            "miktar": item.get("miktar"),
            "birim": item.get("birim"),
        })
    yuk.save()
    return True

@frappe.whitelist()
def has_yuk_permission(doc, ptype, user):
    """Yük izin kontrolü (B2B)"""
    if frappe.has_permission("Lojinet Yuk", ptype="read", user=user):
        return True
    
    customer = frappe.db.get_value("User", user, "lojinet_customer")
    if customer and doc.musteri == customer:
        return True
    return False

# ============== SEFER İŞLEMLERİ ==============
@frappe.whitelist()
def add_yuk_to_sefer(sefer_name, yuk_list):
    """Sefere toplu yük ekleme"""
    import json
    sefer = frappe.get_doc("Lojinet Sefer", sefer_name)
    yukler = json.loads(yuk_list) if isinstance(yuk_list, str) else yuk_list
    
    for yuk_name in yukler:
        yuk = frappe.get_doc("Lojinet Yuk", yuk_name)
        sefer.append("sefer_yukleri", {
            "yuk": yuk.name,
            "irsaliye_no": yuk.musteri_irsaliye_no
        })
        yuk.sefer = sefer.name
        yuk.save(ignore_permissions=True)
    
    sefer.save()
    frappe.msgprint(f"{len(yukler)} adet yük sefere eklendi")
    return True

# ============== FATURA İŞLEMLERİ ==============
@frappe.whitelist()
def create_invoice_from_yuk(yuk_name):
    """Yükten fatura oluştur"""
    yuk = frappe.get_doc("Lojinet Yuk", yuk_name)
    
    invoice = frappe.get_doc({
        "doctype": "Sales Invoice",
        "customer": yuk.musteri,
        "posting_date": today(),
        "due_date": add_days(today(), 30)
    })
    
    for price in yuk.yuk_fiyatlari:
        if price.fatura_durumu == "Faturalanmadı":
            invoice.append("items", {
                "item_code": "NAKLİYE",
                "description": price.masraf_kalemi,
                "qty": 1,
                "rate": price.tutar
            })
    
    invoice.insert()
    
    for price in yuk.yuk_fiyatlari:
        price.fatura_durumu = "Faturalandı"
        price.fatura_no = invoice.name
        price.fatura_tarihi = today()
    yuk.save()
    
    return invoice.name

@frappe.whitelist()
def create_bulk_invoice(customer, from_date=None, to_date=None):
    """Cariye toplu fatura oluştur - Faturalanmamış tüm yükler"""
    if not to_date:
        to_date = today()
    if not from_date:
        from_date = add_days(today(), -30)
    
    # Faturalanmamış yükleri bul
    yukler = frappe.db.sql("""
        SELECT DISTINCT y.name, y.musteri_irsaliye_no
        FROM `tabLojinet Yuk` y
        JOIN `tabLojinet Yuk Fiyat` yf ON yf.parent = y.name
        WHERE y.musteri = %(customer)s
        AND y.islem_tarihi BETWEEN %(from_date)s AND %(to_date)s
        AND yf.fatura_durumu = 'Faturalanmadı'
        AND y.docstatus = 1
        ORDER BY y.islem_tarihi
    """, {"customer": customer, "from_date": from_date, "to_date": to_date}, as_dict=True)
    
    if not yukler:
        frappe.throw("Faturalanmamış yük bulunamadı")
    
    # Tüm faturalanmamış tutarları topla
    total_amount = 0
    yuk_list = []
    
    for yuk_item in yukler:
        yuk = frappe.get_doc("Lojinet Yuk", yuk_item.name)
        yuk_amount = 0
        
        for price in yuk.yuk_fiyatlari:
            if price.fatura_durumu == "Faturalanmadı":
                yuk_amount += flt(price.tutar)
        
        if yuk_amount > 0:
            total_amount += yuk_amount
            yuk_list.append({
                "yuk": yuk.name,
                "irsaliye_no": yuk.musteri_irsaliye_no,
                "tutar": yuk_amount
            })
    
    # Tek fatura oluştur
    invoice = frappe.get_doc({
        "doctype": "Sales Invoice",
        "customer": customer,
        "posting_date": today(),
        "due_date": add_days(today(), 30)
    })
    
    # Tek kalem: Nakliye Bedeli
    invoice.append("items", {
        "item_code": "NAKLİYE",
        "description": f"Nakliye Bedeli - {len(yuk_list)} Adet Yük",
        "qty": 1,
        "rate": total_amount
    })
    
    invoice.insert()
    invoice.submit()
    
    # Tüm yüklerin fiyatlarını güncelle
    for yuk_item in yuk_list:
        yuk = frappe.get_doc("Lojinet Yuk", yuk_item["yuk"])
        for price in yuk.yuk_fiyatlari:
            if price.fatura_durumu == "Faturalanmadı":
                price.fatura_durumu = "Faturalandı"
                price.fatura_no = invoice.name
                price.fatura_tarihi = today()
        yuk.save()
    
    return {
        "fatura_no": invoice.name,
        "yuk_sayisi": len(yuk_list),
        "toplam_tutar": total_amount,
        "yukler": yuk_list
    }

@frappe.whitelist()
def get_invoice_detail_report(invoice_no):
    """Fatura detay raporu - Yük bazında"""
    invoice = frappe.get_doc("Sales Invoice", invoice_no)
    
    # Bu faturaya ait yükleri bul
    yukler = frappe.db.sql("""
        SELECT 
            y.name as yuk_no,
            y.musteri_irsaliye_no,
            y.islem_tarihi,
            yf.masraf_kalemi,
            yf.tutar,
            yf.fatura_no,
            yf.fatura_tarihi
        FROM `tabLojinet Yuk` y
        JOIN `tabLojinet Yuk Fiyat` yf ON yf.parent = y.name
        WHERE yf.fatura_no = %(invoice_no)s
        ORDER BY y.islem_tarihi, y.name
    """, {"invoice_no": invoice_no}, as_dict=True)
    
    # Yük kalemlerini ekle
    for yuk in yukler:
        yuk["kalemler"] = frappe.db.sql("""
            SELECT 
                yd.stok_kodu,
                yd.miktar,
                yd.birim,
                yd.desi
            FROM `tabLojinet Yuk Detay` yd
            WHERE yd.parent = %(yuk_no)s
        """, {"yuk_no": yuk.yuk_no}, as_dict=True)
    
    return {
        "fatura_no": invoice.name,
        "fatura_tarihi": invoice.posting_date,
        "musteri": invoice.customer,
        "toplam_tutar": invoice.grand_total,
        "yukler": yukler
    }

@frappe.whitelist()
def export_invoice_detail_excel(invoice_no):
    """Fatura detay raporunu Excel olarak export et"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO
    import base64
    
    data = get_invoice_detail_report(invoice_no)
    
    # Excel workbook oluştur
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fatura Detay"
    
    # Başlık
    ws['A1'] = "FATURA DETAY RAPORU"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:H1')
    
    ws['A2'] = f"Fatura No: {data['fatura_no']}"
    ws['A3'] = f"Fatura Tarihi: {data['fatura_tarihi']}"
    ws['A4'] = f"Müşteri: {data['musteri']}"
    ws['A5'] = f"Toplam Tutar: {data['toplam_tutar']} TL"
    
    # Tablo başlıkları
    headers = ["Yük No", "İrsaliye No", "Tarih", "Stok Kodu", "Miktar", "Birim", "Desi", "Tutar"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    
    # Veriler
    row = 8
    for yuk in data['yukler']:
        for kalem in yuk.get('kalemler', []):
            ws.cell(row=row, column=1).value = yuk['yuk_no']
            ws.cell(row=row, column=2).value = yuk['musteri_irsaliye_no']
            ws.cell(row=row, column=3).value = str(yuk['islem_tarihi'])
            ws.cell(row=row, column=4).value = kalem['stok_kodu']
            ws.cell(row=row, column=5).value = kalem['miktar']
            ws.cell(row=row, column=6).value = kalem['birim']
            ws.cell(row=row, column=7).value = kalem.get('desi', '')
            ws.cell(row=row, column=8).value = yuk['tutar']
            row += 1
    
    # Dosyayı kaydet
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Base64 encode
    file_data = base64.b64encode(output.read()).decode('utf-8')
    
    return {
        "filename": f"Fatura_Detay_{invoice_no}.xlsx",
        "file_data": file_data
    }

@frappe.whitelist()
def send_invoice_detail_email(invoice_no, recipients=None):
    """Fatura detay raporunu mail olarak gönder"""
    data = get_invoice_detail_report(invoice_no)
    excel_data = export_invoice_detail_excel(invoice_no)
    
    if not recipients:
        customer = data['musteri']
        cari = frappe.get_doc("Lojinet Cari", customer)
        recipients = [cari.email] if cari.email else []
    
    # HTML tablosu oluştur
    html_content = f"""
    <h3>Fatura Detay Raporu</h3>
    <p><strong>Fatura No:</strong> {data['fatura_no']}</p>
    <p><strong>Fatura Tarihi:</strong> {data['fatura_tarihi']}</p>
    <p><strong>Toplam Tutar:</strong> {data['toplam_tutar']} TL</p>
    
    <table border="1" cellpadding="5" cellspacing="0">
        <tr style="background-color: #CCE5FF;">
            <th>Yük No</th>
            <th>İrsaliye No</th>
            <th>Tarih</th>
            <th>Tutar</th>
        </tr>
    """
    
    for yuk in data['yukler']:
        html_content += f"""
        <tr>
            <td>{yuk['yuk_no']}</td>
            <td>{yuk['musteri_irsaliye_no']}</td>
            <td>{yuk['islem_tarihi']}</td>
            <td>{yuk['tutar']} TL</td>
        </tr>
        """
    
    html_content += "</table><p>Detaylı rapor için ekteki Excel dosyasını inceleyiniz.</p>"
    
    # Mail gönder
    frappe.sendmail(
        recipients=recipients,
        subject=f"Fatura Detay Raporu - {data['fatura_no']}",
        message=html_content,
        attachments=[{
            "fname": excel_data['filename'],
            "fcontent": excel_data['file_data']
        }]
    )
    
    return True

# ============== DESTEK BİLETİ İŞLEMLERİ ==============
@frappe.whitelist()
def auto_close_tickets():
    """12 saat yanıt alınamayan biletleri kapat"""
    from frappe.utils import add_to_date
    
    cutoff_time = add_to_date(now(), hours=-12)
    
    tickets = frappe.get_all("Lojinet Destek Bileti",
        filters={"durum": ["!=", "Kapalı"], "modified": ["<", cutoff_time]},
        fields=["name"]
    )
    
    for ticket in tickets:
        doc = frappe.get_doc("Lojinet Destek Bileti", ticket.name)
        doc.durum = "Kapalı"
        doc.save(ignore_permissions=True)

@frappe.whitelist()
def has_ticket_permission(doc, ptype, user):
    """Destek bileti izin kontrolü"""
    if frappe.has_permission("Lojinet Destek Bileti", ptype="read", user=user):
        return True
    
    customer = frappe.db.get_value("User", user, "lojinet_customer")
    if customer and doc.musteri == customer:
        return True
    return False

# ============== ARAÇ İŞLEMLERİ ==============
@frappe.whitelist()
def check_insurance_expiry():
    """Sigorta bitiş tarihlerini kontrol et"""
    from frappe.utils import getdate, add_days
    
    upcoming_date = add_days(today(), 30)
    
    vehicles = frappe.get_all("Lojinet Arac",
        filters={"sigorta_bitis_tarihi": ["between", [today(), upcoming_date]]},
        fields=["name", "plaka", "sigorta_bitis_tarihi", "arac_sahibi"]
    )
    
    for vehicle in vehicles:
        # Mail gönder veya bildirim oluştur
        pass

# ============== STOK İŞLEMLERİ ==============
@frappe.whitelist()
def get_available_stock(item_code, customer):
    """Müşteriye ait boştaki stok"""
    sql = """
        SELECT SUM(mkd.miktar) as toplam
        FROM `tabLojinet Mal Kabul Detay` mkd
        JOIN `tabLojinet Mal Kabul` mk ON mk.name = mkd.parent
        WHERE mkd.stok_kodu = %s
        AND mk.cari = %s
        AND mk.docstatus = 1
    """
    result = frappe.db.sql(sql, (item_code, customer), as_dict=True)
    return result[0].toplam if result else 0

# ============== BOOT SESSION ==============
def boot_session(bootinfo):
    """Boot session'a özel bilgiler ekle"""
    bootinfo.lojinet_version = "1.0.0"
    
    if frappe.session.user != "Guest":
        customer = frappe.db.get_value("User", frappe.session.user, "lojinet_customer")
        if customer:
            bootinfo.lojinet_customer = customer

# ============== RAPORLAMA ==============
@frappe.whitelist()
def get_cari_bakiye(cari):
    """Cari bakiye hesapla"""
    # Toplam satışlar
    total_sales = frappe.db.sql("""
        SELECT SUM(yf.tutar)
        FROM `tabLojinet Yuk Fiyat` yf
        JOIN `tabLojinet Yuk` y ON y.name = yf.parent
        WHERE y.musteri = %s AND yf.fatura_durumu = 'Faturalandı'
    """, cari)[0][0] or 0
    
    # Toplam ödemeler
    total_payments = frappe.db.sql("""
        SELECT SUM(tutar)
        FROM `tabLojinet Odeme Tahsilat`
        WHERE cari = %s AND islem_tipi = 'Tahsilat'
    """, cari)[0][0] or 0
    
    return {
        "borc": total_sales,
        "alacak": total_payments,
        "bakiye": total_sales - total_payments
    }

# ============== MAIL GÖNDERME ==============
@frappe.whitelist()
def send_mail_with_pdf(doctype, docname, recipients=None):
    """PDF ile mail gönder"""
    doc = frappe.get_doc(doctype, docname)
    
    if not recipients:
        if doctype == "Lojinet Mal Kabul":
            cari = frappe.get_doc("Lojinet Cari", doc.cari)
            recipients = [cari.email] if cari.email else []
    
    pdf = frappe.attach_print(doctype, docname)
    
    frappe.sendmail(
        recipients=recipients,
        subject=f"{doctype} - {docname}",
        message=f"<p>{doctype} belgesi ektedir.</p>",
        attachments=[pdf]
    )
    return True

# ============== SAAS İŞLEMLERİ ==============
@frappe.whitelist()
def create_saas_customer(firma_adi, paket, donem="Aylık"):
    """SaaS müşteri oluştur"""
    from dateutil.relativedelta import relativedelta
    
    musteri = frappe.get_doc({
        "doctype": "Lojinet Saas Musteri",
        "firma_adi": firma_adi,
        "paket": paket,
        "baslangic_tarihi": today(),
        "bitis_tarihi": add_days(today(), 30 if donem == "Aylık" else 365),
        "durum": "Deneme"
    })
    musteri.insert()
    return musteri.name
